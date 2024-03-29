from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from .models import *
from datetime import datetime, date, time
from django.http import Http404
from .forms import IssuesForm, IssuesEditForm, ReportForm, Search_for_Number, StatisticForm, StatisticDownForm
from django.views.i18n import *
from django.contrib.auth.models import User
from login.models import *
from django.http import HttpResponse
from django.core import serializers
from django.conf import settings
from django.http import JsonResponse
import json
import re
import openpyxl
import os

global mod_list
mod_list = []
type_list = []
inv_list = []


def index(request):
    issue = issues.objects.order_by("-change_date")
    user_warn = request.user
    if user_warn.is_authenticated:
        current_user = UserProfile.objects.get(user=user_warn)
        current_user_role = str(current_user.id_state)
        return render(request, 'manager/index.html', {'issues': issue, 'current_user_role': current_user_role})

    else:

        return redirect('/login/')


# функция для получения переменной с типом оборудования из ajax post-запроса
def get_type(request):
    type_id = request.POST.get('cat_id')

    if type_id != None:

        type_list.insert(0, type_id)

        t = type_list[0]

        # type_list.clear()

        return t

    else:

        t = type_list[0]
        type_list.clear()
        return t


# функция для получения переменной с именем модели из ajax post-запроса
def get_model(request):
    mod_id = request.POST.get('model')

    if mod_id != None:

        mod_list.insert(0, mod_id)

        m = mod_list[0]

        # mod_list.clear()

        return m

    else:

        m = mod_list[0]
        mod_list.clear()
        return m


# функция для получения переменной с инвентарным номером оборудования из ajax post-запроса
def get_inventory(request):
    inv = request.POST.get('num')

    if inv != None:
        inv_list.insert(0, inv)
        i = inv_list[0]
        # inv_list.clear()
        return i
    else:
        i = inv_list[0]
        inv_list.clear()
        return i


# view создания инцидента
def create_issue(request):
    user_warn = request.user
    if user_warn.is_authenticated:
        current_user = UserProfile.objects.get(user=user_warn)
        current_user_role = str(current_user.id_state)
        new_issues = issues()
        number = issues.objects.values('number_issue').order_by().last()
        if number is None:
            num_view = 0
        else:
            num_view = number['number_issue'] + 1
        form = IssuesForm()
        # заполняем поля для выбора оборудования через ajax-запрос
        if request.is_ajax() == True:
            if request.method == "GET":
                space = request.GET.get('space')
                if space != None:
                    space_id = workspace.objects.get(name=space)
                    type_inventory = str(equipment.objects.get_name(space_id.id))
                    dist_work = re.findall(r'[\d\w\s,-]+', type_inventory)
                    return HttpResponse(dist_work)
                cat_id = request.GET.get('name')
                if cat_id != None:
                    model = str(equipment.objects.get_model(cat_id))
                    dist = re.findall(r'[\d\w,]+', model)
                    return HttpResponse(dist)
                mod = request.GET.get('model')
                if mod != None:
                    inventory = str(equipment.objects.get_inventory(mod))
                    dist_inventory = re.findall(r'[\d\w,]+', inventory)
                    return HttpResponse(dist_inventory)
                inv_num = request.GET.get('send')
                # добавить обнуление переменных
                if inv_num != None:
                    sender = str(equipment.objects.check_inventory(inv_num))
                    send_dist = re.findall(r'[\d\w,]+', sender)
                    result = "Оборудование успешно выбрано, нажмите кнопку ""Сохранить"""
                    return HttpResponse(result)

        if request.method == "POST":
            form = IssuesForm(request.POST)
            if request.is_ajax() == True:
                # отправляем значения в списки
                type_id = get_type(request)
                mod_id = get_model(request)
                inv_num_id = get_inventory(request)
                if type_id == "" or mod_id == "" or inv_num_id == "":
                    result = "Выбор не сохранен.Выберите все необходимые поля"
                else:
                    result = "Выбор сохранен успешно №%s, тип %s, модель %s" % (inv_num_id, type_id, mod_id)

                return HttpResponse(result)
            # забираем значения из списков
            name = get_type(request)
            model = get_model(request)
            inventory = get_inventory(request)
            if form.is_valid():
                new_issues = form.save(commit=False)
                new_issues.equipment_name = equipment.objects.filter(name=name).last()
                new_issues.equipment_model = equipment.objects.filter(model=model).last()
                new_issues.equipment_inventory = equipment.objects.get(inventory_number=inventory)
                new_issues.number_issue = num_view
                new_issues.number_history = "1"
                new_issues.user_edit = request.user
                new_issues.change_date = timezone.now()
                new_issues.save()
                return redirect('/index/')

            else:
                return HttpResponse("Форма невалидна")
        return render(request, 'manager/create_issue.html',
                      {'form': form, 'num_view': num_view, 'current_user_role': current_user_role})
    else:
        return redirect('/login/')


# представление просмотра информации по инциденту
def view_issue(request, number):
    user_warn = request.user

    if user_warn.is_authenticated:

        current_user = UserProfile.objects.get(user=user_warn)

        current_user_role = str(current_user.id_state)

        issue = get_object_or_404(issues, number_issue=number)

        model = str(equipment.objects.values('model').filter(pk=issue.equipment_model_id))

        inventory = str(equipment.objects.values('inventory_number').filter(pk=issue.equipment_model_id))

        model = re.sub(r'QuerySet', ' ', model)

        model = re.sub(r'model', ' ', model)

        model = str(re.findall(r'([А-Я]+|[A-Z]+|[0-9]+)', model))

        model = re.sub("(\['|\'])", ' ', model)

        model = re.sub("', '", '', model)

        inventory = re.sub(r'QuerySet', ' ', inventory)

        inventory = re.sub(r'inventory_number', ' ', inventory)

        inventory = str(re.findall(r'([0-9]+)', inventory))

        inventory = re.sub("(\['|\'])", ' ', inventory)

        return render(request, 'manager/issue.html',
                      {'issue': issue, 'model': model, 'inventory': inventory, 'current_user_role': current_user_role})

    else:

        return redirect('/login/')


# представление дл редактирования инцидента
def issue_edit(request, number):
    user_warn = request.user

    if user_warn.is_authenticated:

        current_user = UserProfile.objects.get(user=user_warn)

        current_user_role = str(current_user.id_state)

        issue = issues.objects.get(number_issue=number)

        issue = get_object_or_404(issues, number_issue=number)
        # вытягивает данные по тем полям, которых не будет в форме, иначе не сохраним модель
        workspace = issue.workspace
        number = issue.number_issue
        creator = issue.creator_id
        inv_number = issue.equipment_inventory_id
        model_id = issue.equipment_model_id
        type_eq = issue.equipment_name_id
        number_history = issue.number_history + 1
        name_id = int(issue.equipment_name_id)
        # получаем результат запроса для вывода модели и №
        # получаем модель
        model = equipment.objects.filter(pk=name_id)
        model_result = str(model.values_list('model'))
        dist_model = re.sub(r'QuerySet', ' ', model_result)
        dist_model_result = str(re.findall(r'([А-Я]+|[A-Z]+|[0-9]+)', dist_model))
        dist_model_result = re.sub("(\['|\'])", ' ', dist_model_result)
        dist_model_result = re.sub("', '", '', dist_model_result)
        # получаем №
        inv_num = str(model.values_list('inventory_number'))
        dist_num = re.sub(r'QuerySet', ' ', inv_num)
        dist_num_result = str(re.findall(r'([А-Я]+|[A-Z]+|[0-9]+)', dist_num))
        dist_num_result = re.sub("(\['|\'])", ' ', dist_num_result)
        form = IssuesEditForm(instance=issue)
        if request.method == "POST":
            form = IssuesEditForm(request.POST)

            if form.is_valid():
                issue = form.save(commit=False)
                issue.workspace = workspace
                issue.creator_id = creator
                issue.number_issue = number
                issue.equipment_inventory_id = inv_number
                issue.equipment_model_id = model_id
                issue.equipment_name_id = type_eq
                issue.number_history = number_history
                issue.user_edit = request.user
                issue.change_date = timezone.now()
                issue.save()
                if not issue.save:
                    return HttpResponse("База данных недоступна. Попробуйте позднее")
            else:
                result_save_unsucces = "Форма невалидна"
                return HttpResponse(result_save_unsucces)
        if request.is_ajax() == True:
            issue1 = get_object_or_404(issues, number_issue=number)
            if issue1.number_history == (number_history - 1):
                result_save_succes = "Изменения сохранены "
                return HttpResponse(result_save_succes)
            else:
                result_save_unsucces = "Сохранение не произошло"
                return HttpResponse(result_save_unsucces)

        return render(request, 'manager/issue_edit.html',
                      {'form': form, 'issue': issue, 'dist_model': dist_model_result,
                       'dist_num_result': dist_num_result, 'current_user_role': current_user_role})

    else:

        return redirect('/login/')


# список инцидентов, созданных или находящихся в работе у пользователя
def view_issues_user(request):
    user_warn = request.user
    current_user = UserProfile.objects.get(user=request.user)
    issues_user = current_user
    current_user_role = str(current_user.id_state)
    if user_warn.is_authenticated:

        if current_user_role == "Координатор":

            issue_user = issues.objects.filter(coordinator=current_user)

        elif current_user_role == "Инициатор":

            issue_user = issues.objects.filter(creator=current_user)

        elif current_user_role == "Исполнитель":

            issue_user = issues.objects.filter(executor=current_user)

        elif current_user_role == "Администратор":

            issue_user = issues.objects.filter(administrator=current_user)

        return render(request, 'manager/issues_user.html',
                      {'issue_user': issue_user, 'current_user_role': current_user_role})
    else:

        return redirect('/login/')


# список инцидентов, созданных или находящихся в работе у группы пользователя
def view_issues_user_groups(request):
    user_warn = request.user
    current_user = UserProfile.objects.get(user=request.user)
    issues_user = current_user
    if user_warn.is_authenticated:
        if current_user.group_of_work_id != None:
            current_user_group = str(current_user.group_of_work_id)
        else:
            current_user_group = current_user.group_of_work_id

        issues_groups = issues.objects.filter(groups_of_work=current_user_group)

        current_user_role = str(current_user.id_state)

        return render(request, 'manager/issues_group.html',
                      {'issues_groups': issues_groups, 'current_user_role': current_user_role})
    else:

        return redirect('/login/')


# список инцидентов, ожидающих ответа от пользователя
def view_issues_user_wait(request):
    user_warn = request.user
    if user_warn.is_authenticated:

        current_user = UserProfile.objects.get(user=request.user)
        issues_user = current_user
        current_user_role = str(current_user.id_state)

        if current_user_role == "Координатор":

            issue_wait = issues.objects.filter(coordinator=current_user, current_status=1)

        elif current_user_role == "Инициатор":

            issue_wait = issues.objects.filter(creator=current_user, current_status=(5, 6, 7))

        elif current_user_role == "Исполнитель":

            issue_wait = issues.objects.filter(executor=current_user, current_status=2)

        return render(request, 'manager/issue_user_wait.html',
                      {'issue_wait': issue_wait, 'current_user_role': current_user_role})

    else:

        return redirect('/login/')


def view_reports(request):
    user_warn = request.user
    if user_warn.is_authenticated:
        form = ReportForm()
        current_user = UserProfile.objects.get(user=request.user)
        issues_user = current_user
        current_user_role = str(current_user.id_state)
        if request.is_ajax() == True:
            form_send = request.POST.get('data')
            start_period = re.findall(r'(start_period=[0-9.]+)', form_send)
            start_period = str(start_period)
            start_period_result_1 = str(re.findall(r'([0-9.]+)', start_period))
            start_period_result_2 = str(re.sub("(\['|\'])", '', start_period_result_1))
            end_period = re.findall(r'(end_period=[0-9.]+)', form_send)
            end_period = str(end_period)
            end_period_result_1 = str(re.findall(r'([0-9.]+)', end_period))
            end_period_result_2 = str(re.sub("(\['|\'])", '', end_period_result_1))
            format_report = request.POST.get('format')
            start_period_result = datetime.strptime(str(start_period_result_2), "%d.%m.%Y")
            end_period_result = datetime.strptime(str(end_period_result_2), "%d.%m.%Y")
            # блок работы с Excel
            filename = "Отчёт_" + str(start_period_result_2) + "_" + str(end_period_result_2) + ".xls"
            report_book = openpyxl.Workbook()
            report_sheet = report_book.active
            report_sheet.title = 'Инциденты'
            # report_sheet = report_book.create_sheet("Инциденты")
            report_sheet.cell(row=1, column=1).value = "№ инцидента"
            report_sheet.cell(row=1, column=2).value = "Уровень"
            report_sheet.cell(row=1, column=3).value = "Статус"
            report_sheet.cell(row=1, column=4).value = "Дата начала простоя"
            report_sheet.cell(row=1, column=5).value = "Время начала простоя"
            report_sheet.cell(row=1, column=6).value = "Дата создания инцидента"
            report_sheet.cell(row=1, column=7).value = "Время создания инцидента"
            report_sheet.cell(row=1, column=8).value = "Дата окончания простоя"
            report_sheet.cell(row=1, column=9).value = "Время начала простоя"
            report_sheet.cell(row=1, column=10).value = "Поздразделение"
            report_sheet.cell(row=1, column=11).value = "Наименование"
            report_sheet.cell(row=1, column=12).value = "Модель"
            report_sheet.cell(row=1, column=13).value = "Инвентарный №"
            report_sheet.cell(row=1, column=14).value = "Координатор"
            i = 1

            j = 1
            report = equipment.objects.get_report(start_period_result, end_period_result)
            for rep in range(len(report)):
                i = i + 1
                j = 1
                for r in report[rep]:
                    report_sheet.cell(row=i, column=j).value = r
                    j = j + 1

            page_path = os.path.join(settings.MEDIA_ROOT, 'reports/%s' % (filename))
            page_save = 'reports/%s' % (filename)
            report_book.save(page_path)
            rep = report_book
            new_report = Reports(creator=request.user, report=page_path)
            new_report.save()
            return HttpResponse(page_save)
        return render(request, 'manager/report_page.html', {'form': form, 'current_user_role': current_user_role})

    else:

        return redirect('/login/')


def find_issues(request):
    user_warn = request.user
    if user_warn.is_authenticated:
        form = Search_for_Number()
        current_user = UserProfile.objects.get(user=request.user)
        issues_user = current_user
        current_user_role = str(current_user.id_state)
        if request.is_ajax() == True:
            find_number = int(request.GET.get('num'))
            issue = equipment.objects.get_number(find_number)
            # issue_result = str(re.sub("(['|'])", '', issue))
            # if issue != None:
            issue_list_result = []
            # обработка результата для разрешения точек и запятых в полях с описанием
            for r in range(len(issue)):
                for i in issue[r]:
                    issue_list_result.append("\'" + str(i) + "\'")

            return HttpResponse(issue_list_result)

        return render(request, 'manager/find_issues.html', {'form': form, 'current_user_role': current_user_role})
    else:

        return redirect('/login/')


def find_content(request):
    user_warn = request.user
    if user_warn.is_authenticated:
        form = Search_for_Number()
        current_user = UserProfile.objects.get(user=request.user)
        issues_user = current_user
        current_user_role = str(current_user.id_state)
        if request.is_ajax() == True:
            find_content = request.GET.get('num')
            find_content_2 = str(re.sub("'", '', find_content))
            issue = equipment.objects.get_source(find_content)
            issue_list_result = []
            # обработка результата для разрешения точек и запятых в полях с описанием
            for r in range(len(issue)):
                for i in issue[r]:
                    issue_list_result.append("\'" + str(i) + "\'")

            return HttpResponse(issue_list_result)

        return render(request, 'manager/find_content.html', {'form': form, 'current_user_role': current_user_role})
    else:

        return redirect('/login/')


def view_statistic_downtime(request):
    user_warn = request.user
    if user_warn.is_authenticated:
        current_user = UserProfile.objects.get(user=request.user)
        issues_user = current_user
        current_user_role = str(current_user.id_state)
        form = StatisticDownForm()
        if request.is_ajax() == True:
            if request.method == "GET":
                space = request.GET.get('space')
                if space != None:
                    space_id = workspace.objects.get(name=space)
                    type_inventory = str(equipment.objects.get_name(space_id.id))
                    dist_work = re.findall(r'[\d\w\s,-]+', type_inventory)
                    return HttpResponse(dist_work)
                cat_id = request.GET.get('name')
                if cat_id != None:
                    model = str(equipment.objects.get_model(cat_id))
                    dist_model = re.findall(r'[\d\w,]+', model)
                    return HttpResponse(dist_model)
                mod = request.GET.get('model')
                if mod != None:
                    inventory = str(equipment.objects.get_inventory(mod))
                    dist_inventory = re.findall(r'[\d\w,]+', inventory)
                    return HttpResponse(dist_inventory)
                inv_num = request.GET.get('send')
                if inv_num != None:
                    sender = str(equipment.objects.check_inventory(inv_num))
                    send_dist = re.findall(r'[\d\w,]+', sender)
                    result = "Оборудование успешно выбрано, нажмите кнопку ""Отправить"" для получения результата"
                    return HttpResponse(result)
            form_send = request.POST.get('data')
            start_period = re.findall(r'(start_period=[0-9.]+)', form_send)
            start_period = str(start_period)
            start_period_result_1 = str(re.findall(r'([0-9.]+)', start_period))
            start_period_result_2 = str(re.sub("(\['|\'])", '', start_period_result_1))
            end_period = re.findall(r'(end_period=[0-9.]+)', form_send)
            end_period = str(end_period)
            end_period_result_1 = str(re.findall(r'([0-9.]+)', end_period))
            end_period_result_2 = str(re.sub("(\['|\'])", '', end_period_result_1))
            start_period_result = datetime.strptime(str(start_period_result_2), "%d.%m.%Y")
            end_period_result = datetime.strptime(str(end_period_result_2), "%d.%m.%Y")
            work_space = str(re.findall(r'(workspace=[0-9]+)', form_send))
            work_space = str(re.findall(r'([0-9.]+)', work_space))
            work_space = str(re.sub("(\['|\'])", '', work_space))
            model = request.POST.get('model')
            inv = request.POST.get('inv')
            issue = equipment.objects.get_stat_downtime(start_period_result, end_period_result, work_space, model, inv)
            list_issue = []
            for i in issue:
                for n in i:
                    n = [str(n)]
                    list_issue.append(n)
            return HttpResponse(str(list_issue))

        return render(request, 'manager/view_statistic_downtime.html',
                      {'form': form, 'current_user_role': current_user_role})
    else:

        return redirect('/login/')


def view_statistic_issues(request):
    user_warn = request.user
    if user_warn.is_authenticated:
        current_user = UserProfile.objects.get(user=request.user)
        issues_user = current_user
        current_user_role = str(current_user.id_state)
        form = StatisticForm()
        if request.is_ajax() == True:
            form_send = request.POST.get('data')
            start_period = re.findall(r'(start_period=[0-9.]+)', form_send)
            start_period = str(start_period)
            start_period_result_1 = str(re.findall(r'([0-9.]+)', start_period))
            start_period_result_2 = str(re.sub("(\['|\'])", '', start_period_result_1))
            end_period = re.findall(r'(end_period=[0-9.]+)', form_send)
            end_period = str(end_period)
            end_period_result_1 = str(re.findall(r'([0-9.]+)', end_period))
            end_period_result_2 = str(re.sub("(\['|\'])", '', end_period_result_1))
            start_period_result = datetime.strptime(str(start_period_result_2), "%d.%m.%Y")
            end_period_result = datetime.strptime(str(end_period_result_2), "%d.%m.%Y")
            stat_issue = equipment.objects.get_stat_level_issue(start_period_result, end_period_result)
            stat_status = equipment.objects.get_stat_status_issue(start_period_result, end_period_result)
            dist_stat = dict(stat_issue + stat_status)
            dist_stat_result = json.dumps(dist_stat)
            return HttpResponse(dist_stat_result)

        return render(request, 'manager/view_statistic_issues.html',
                      {'form': form, 'current_user_role': current_user_role})
    else:

        return redirect('/login/')


def issue_history(request, number):
    user_warn = request.user

    if user_warn.is_authenticated:

        if request.method == "GET":
            current_user = UserProfile.objects.get(user=user_warn)

            current_user_role = str(current_user.id_state)

            issue = equipment.objects.history_issues(number)

            return render(request, 'manager/issue_history.html',
                          {'issues': issue, 'current_user_role': current_user_role})

    else:

        return redirect('/login/')
